#encoding=utf-8
#Programmer：411147372 馬芯瑜
#Date ：2024/05/23
#彰化縣建照
#pip install openpyxl
#pip install selenium
#pip install ddddocr
#pip install webdriver-manager
#pip install pillow

# 使用 webdriver 管理器4.0.1
# 版本selenium 4
#python 3.11
import re
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import requests
# 獲取預設的工作表
# 設定合併儲存格的文字居中樣式
import openpyxl
from openpyxl.styles import Alignment

alignment = Alignment(horizontal="center", vertical="center", wrapText=True)  # 置中換行

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "建築執照"
sheet2 = workbook.create_sheet("使用執照")


sheet1.merge_cells("A1:A2")
merged_cell = sheet1["A1"]
merged_cell.value = "縣市"
merged_cell.alignment = alignment

#sheet1.column_dimensions["B:C"].width = 20
sheet1.merge_cells("B1:C2")
merged_cell = sheet1["B1"]
merged_cell.value = "建築執照"
merged_cell.alignment = alignment

sheet1.merge_cells("D1:D2")
merged_cell = sheet1["D1"]
merged_cell.value = "起造人"
merged_cell.alignment = alignment

sheet1.merge_cells("E1:F2")
merged_cell = sheet1["E1"]
merged_cell.value = "設計人"
merged_cell.alignment = alignment

sheet1.merge_cells("G1:H2")
merged_cell = sheet1["G1"]
merged_cell.value = "承造人"
merged_cell.alignment = alignment

sheet1.merge_cells("I1:I2")
merged_cell = sheet1["I1"]
merged_cell.value = "地號"
merged_cell.alignment = alignment

sheet1.merge_cells("J1:J2")
merged_cell = sheet1["J1"]
merged_cell.value = "使用分區"
merged_cell.alignment = alignment

sheet1.merge_cells("K1:K2")
merged_cell = sheet1["K1"]
merged_cell.value = "層棟戶數"
merged_cell.alignment = alignment

sheet1.merge_cells("L1:L2")
merged_cell = sheet1["L1"]
merged_cell.value = "總樓地板面積(㎡)"
merged_cell.alignment = alignment

sheet1.merge_cells("M1:O1")
merged_cell = sheet1["M1"]
merged_cell.value = "合計"
merged_cell.alignment = alignment

merged_cell = sheet1["M2"]
merged_cell.value = "個案數"
merged_cell.alignment = alignment

merged_cell = sheet1["N2"]
merged_cell.value = "戶數"
merged_cell.alignment = alignment

merged_cell = sheet1["O2"]
merged_cell.value = "總樓地板面積"
merged_cell.alignment = alignment

#使用執照
sheet2.merge_cells("A1:A2")
merged_cell = sheet2["A1"]
merged_cell.value = "縣市"
merged_cell.alignment = alignment

sheet2.merge_cells("B1:B2")
merged_cell = sheet2["B1"]
merged_cell.value = "使用執照"
merged_cell.alignment = alignment
sheet2.merge_cells("C1:C2")
merged_cell = sheet2["C1"]
merged_cell.value = "建築執照"
merged_cell.alignment = alignment

sheet2.merge_cells("D1:D2")
merged_cell = sheet2["D1"]
merged_cell.value = "起造人"
merged_cell.alignment = alignment

sheet2.merge_cells("E1:F2")
merged_cell = sheet2["E1"]
merged_cell.value = "設計人"
merged_cell.alignment = alignment

sheet2.merge_cells("G1:H2")
merged_cell = sheet2["G1"]
merged_cell.value = "承造人"
merged_cell.alignment = alignment

sheet2.merge_cells("I1:I2")
merged_cell = sheet2["I1"]
merged_cell.value = "地號"
merged_cell.alignment = alignment

sheet2.merge_cells("J1:J2")
merged_cell = sheet2["J1"]
merged_cell.value = "使用分區"
merged_cell.alignment = alignment

sheet2.merge_cells("K1:K2")
merged_cell = sheet2["K1"]
merged_cell.value = "層棟戶數"
merged_cell.alignment = alignment

sheet2.merge_cells("L1:L2")
merged_cell = sheet2["L1"]
merged_cell.value = "總樓地板面積(㎡)"
merged_cell.alignment = alignment

sheet2.merge_cells("M1:O1")
merged_cell = sheet2["M1"]
merged_cell.value = "合計"
merged_cell.alignment = alignment

merged_cell = sheet2["M2"]
merged_cell.value = "個案數"
merged_cell.alignment = alignment

merged_cell = sheet2["N2"]
merged_cell.value = "戶數"
merged_cell.alignment = alignment

merged_cell = sheet2["O2"]
merged_cell.value = "總樓地板面積"
merged_cell.alignment = alignment
sheet1.column_dimensions["B"].width = 12.5
sheet1.column_dimensions["C"].width = 12.5
sheet2.column_dimensions["B"].width = 12.5
sheet2.column_dimensions["C"].width = 12.5
#建立字典d = {"縣市": [建案[],合計[個案數,戶數,總樓地板面積]]}
#'員林巿': [[], [0, 0, 0.0]]
towns1={}
towns2={}
def case():
    # 輸入驗證碼
    # 下載驗證碼檔案image
    driver.save_screenshot('image.png')
    from PIL import Image
    image = Image.open('image.png')
    # 螢幕截圖切割
    image = image.crop((970, 530, 1200, 580))  # left[x],top[y],x+寬,y+高
    # image.show()
    # 驗證碼圖片
    image.save('image.png')
    # 讀取數字
    import ddddocr
    f = open('D:/pycode/test/image.png', mode='rb')  # 圖片路徑,rb=read binary
    img = f.read()
    ocr = ddddocr.DdddOcr()
    result = ocr.classification(img)
    input = driver.find_element("id", "inputCode")
    input.send_keys(result)
    # 查詢
    element = driver.find_element('name', 'send')
    element.click()
    time.sleep(5)

    # 查詢筆數
    # <div class="tabres" style="display: block;">
    tab = driver.find_element(By.CLASS_NAME, 'tabres')
    # 使用 JavaScript 修改按鈕的 CSS 屬性
    driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.opacity = '1';", tab)
    # print(tab.text)
    link = tab.find_element(By.CLASS_NAME, 'ui-jqgrid-bdiv')
    driver.execute_script("arguments[0].style.position = 'relative'; arguments[0].style.opacity = '1';", link)

    td_title = link.find_elements(By.XPATH, '//td[@title]')

    # 提取並打印所有的 title 屬性值
    i = 0
    list = []  # list==每筆資料頁面連結
    for td in td_title:
        i = i + 1
        driver.execute_script("arguments[0].className = 'jqgrow ui-row-ltr active';", td)
        title_value = td.get_attribute('title')
        if i == 2:
            list.append(title_value)
        if i == 6:
            i = -1
    # https://cpami.chcg.gov.tw/bupic/pages/queryInfoAction.do?INDEX_KEY="list"
    for l in list:
        url_n = f"https://cpami.chcg.gov.tw/bupic/pages/queryInfoAction.do?INDEX_KEY={l}"
        driver_n = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))
        response_n = requests.get(url, headers=headers)
        driver_n.get(url_n)
        driver_n.get(url_n)
        # <div class="tableCon">
        # <div class="tit_01">建造執照號碼：</div>
        table = driver_n.find_element(By.CLASS_NAME, 'tableCon')
        lines = table.text.splitlines()
        t = []
        for i in range(0, len(lines)):
            t.append(' '.join(lines[i].split()))
        list_temp = []  # towns1[key][0].append(list_temp)
        for j in range(2, len(t)):
            if "地址" in t[j]:
                key=str(t[j+1][3:6])
        list_temp.append(t[1])
        for j in range(2, len(t)):
            if "建造執照號碼" in t[j]:
                list_temp.append(t[j + 1])
            if "起造人" in t[j]:
                list_temp.append(t[j + 2])
            if "設計人" in t[j]:
                list_temp.append(t[j + 2])
                list_temp.append(t[j + 4])
            if "承造人" in t[j]:
                list_temp.append(t[j + 2])
                list_temp.append(t[j + 4])
            if "地號" in t[j] and "土地" in t[j]:
                list_temp.append(t[j])
            if "使用分區" in t[j]:
                list_temp.append(t[j + 1])
            if "層棟戶數" in t[j]:
                list_temp.append(t[j + 1])
            if "總樓地板面積" in t[j]:
                list_temp.append(t[j + 1])
        if "建造執照" in t[0]:
            if towns1.get(key)==None:
                towns1[key] = [[], [0, 0, 0.0]]
            towns1[key][0].append(list_temp)
            towns1[key][1][0] = towns1[key][1][0] + 1
            for j in range(2, len(t)):
                if "層棟戶數" in t[j]:
                    # 正則表達式r'(\d+)戶'來匹配字串中的數字和"戶"這個詞。(\d+)是一個捕獲組,它捕獲一個或多個數字。
                    pattern = r'(\d+)戶'
                    match = re.search(pattern, t[j + 1])
                    towns1[key][1][1] = towns1[key][1][1] + int(match.group(1))
                if "總樓地板面積" in t[j]:
                    math = float(t[j + 1][0:len(t[j + 1]) - 1])
                    towns1[key][1][2] = towns1[key][1][2] + math
        else:
            if towns2.get(key)==None:
                towns2[key] = [[], [0, 0, 0.0]]
            towns2[key][0].append(list_temp)
            towns2[key][1][0] = towns2[key][1][0] + 1
            for j in range(2, len(t)):
                if "層棟戶數" in t[j]:
                    # 正則表達式r'(\d+)戶'來匹配字串中的數字和"戶"這個詞。(\d+)是一個捕獲組,它捕獲一個或多個數字。
                    pattern = r'(\d+)戶'
                    match = re.search(pattern, t[j + 1])
                    towns2[key][1][1] = towns2[key][1][1] + int(match.group(1))
                if "總樓地板面積" in t[j]:
                    math = float(t[j + 1][0:len(t[j + 1]) - 1])
                    towns2[key][1][2] = towns2[key][1][2] + math
#進入網站
url = "https://cpami.chcg.gov.tw/bupic/preLoginFormAction.do"
cookie="JSESSIONID=BF9CBE193A10CCE423109BD20133A4AA; TS01abd37b=01690d24163471885c3b0a1b04d050b4ce16da38216da06e99f318f657cb3533c934706e64a493d46a9ba4eceb76be979be0bfd7dbcc7ca82a4944a41931d66172ae5e4b65; TS01f4024e=01690d241627261e8fd6038e655cf10da9b503e5fbd6064941823b9cec00bb35a07f728fc827dca7b0fbd3f116ca47c99c514bf905"
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'}

driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))
response = requests.get(url,headers=headers)
driver.get(url)


#取得網站程式碼
#html=response.text
#print(html)
#使用者輸入查找範圍
s=[0,0,0]
ss=s
e=[0,0,0]
ee=e
bool=True

while bool:
    print("輸入查詢起始日期：")
    s[0]=input("選擇年度：")
    ss[0] = s[0]
    s[0]=int(s[0])
    s[1]=input("選擇月份：")
    ss[1] = s[1]
    s[1] = int(s[1])
    """
    if s[1]>12 or s[1]<1:
        print("月份錯誤(1~12月)")
        print("-----------------------------------")
        continue
    """
    s[2]=input("選擇日期：")
    ss[2] = s[2]
    s[2] = int(s[2])
    print("輸入查詢結束日期：")
    e[0]=input("選擇年度：")
    ee[0] = e[0]
    e[0] = int(e[0])
    e[1]=input("選擇月份：")
    ee[1] = e[1]
    e[1] = int(e[1])
    """
    if e[1]>12 or e[1]<1:
        print("月份錯誤(1~12月)")
        print("-----------------------------------")
        continue
    """
    e[2]=input("選擇日期：")
    ee[2]=e[2]
    e[2] = int(e[2])
    print("-----------------------------------")
    """
    if s[0]>e[0]:
        print("年度錯誤(起始<結束)：")
    elif s[0] == e[0]:
        print("日期錯誤(起始<結束)：")
    else:
        bool=False
    """
    bool=False
#用發照日期按鈕
#qtype = '5'<label for="QType5"><span>發照日期</span></label>
radio_button = driver.find_element('id', 'QType5')
#<input type="hidden" id="regdat" name="regdat">
#按鈕被隱藏
# 使用 JavaScript 修改按鈕的 CSS 屬性
driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.opacity = '1';", radio_button)
if radio_button.is_displayed() and radio_button.is_enabled():
    radio_button.click()
else:
    print("Element not visible or interactable")
#輸入查詢日期
bool=True
while bool:
    if s[0]==e[0]:
        bool=False
    input_yy = driver.find_element('id', 'yy')
    input_yy.clear()
    input_yy.send_keys(s[0])
    for i in range(s[1],e[1]+1 if bool==False else 13):
        input_mon = driver.find_element('id', 'mon')
        input_mon.clear()
        input_mon.send_keys(i)
        if i==1 or i==3 or i==5 or i==7 or i==8 or i==10 or i==12 :
            for j in range(s[2],e[2]+1 if s[1]==e[1] else 32):
                input_day = driver.find_element('id', 'day')
                input_day.clear()
                input_day.send_keys(j)
                time.sleep(2)
                case()
        elif i==4 or i==6 or i==9 or i==11 :
            for j in range(s[2],e[2]+1 if s[1]==e[1] else 31):
                input_day = driver.find_element('id', 'day')
                input_day.clear()
                input_day.send_keys(j)
                time.sleep(2)
                case()
        else:
            if s[0]%4==0:
                for j in range(s[2],e[2]+1 if s[1]==e[1] else 30):
                    input_day = driver.find_element('id', 'day')
                    input_day.clear()
                    input_day.send_keys(j)
                    time.sleep(2)
                    case()
            else:
                for j in range(s[2],e[2]+1 if s[1]==e[1] else 29):
                    input_day = driver.find_element('id', 'day')
                    input_day.clear()
                    input_day.send_keys(j)
                    time.sleep(2)
                    case()
        s[2]=1

    s[0]=s[0]+1
    s[1]=0
# 重新寫入排序後的數據
j=int(3)
for i in towns1:#I==key
    merged_cell = sheet1["A" +str(j)]
    merged_cell.value = i
    merged_cell.alignment = alignment
    merged_cell = sheet1["M" +str(j)]
    merged_cell.value = towns1[i][1][0]
    merged_cell.alignment = alignment
    merged_cell = sheet1["N" +str(j)]
    merged_cell.value = towns1[i][1][1]
    merged_cell.alignment = alignment
    merged_cell = sheet1["O" +str(j)]
    merged_cell.value = towns1[i][1][2]
    merged_cell.alignment = alignment
    sheet1.merge_cells("A"+str(j)+":A"+str(j+int(towns1[i][1][0]-1)))
    sheet1.merge_cells("M"+str(j)+":M"+str(j+int(towns1[i][1][0]-1)))
    sheet1.merge_cells("N"+str(j)+":N"+str(j+int(towns1[i][1][0]-1)))
    sheet1.merge_cells("O"+str(j)+":O"+str(j+int(towns1[i][1][0]-1)))
    for k in towns1[i][0]:#k建案
        merged_cell = sheet1["B" +str(j)]
        merged_cell.value = k[0]
        merged_cell.alignment = alignment
        if "起造人" not in k[1]:
            merged_cell = sheet1["C" +str(j)]
            merged_cell.value = k[1]
            merged_cell.alignment = alignment
        else:
            sheet1.merge_cells("B" + str(j) + ":C" + str(j))
        merged_cell = sheet1["D" +str(j)]
        merged_cell.value = k[2]
        merged_cell.alignment = alignment
        merged_cell = sheet1["E" +str(j)]
        merged_cell.value = k[3]
        merged_cell.alignment = alignment
        if "監造人" in k[4]:
            sheet1.merge_cells("E" + str(j) + ":F" + str(j))
        else:
            merged_cell = sheet1["F" +str(j)]
            merged_cell.value = k[4]
            merged_cell.alignment = alignment
        if "營造廠" in k[5]:
            sheet1.merge_cells("G" + str(j) + ":H" + str(j))
            merged_cell = sheet1["G" +str(j)]
            merged_cell.value = "無"
            merged_cell.alignment = alignment
        else:
            merged_cell = sheet1["G" +str(j)]
            merged_cell.value = k[5]
            merged_cell.alignment = alignment
            merged_cell = sheet1["H" +str(j)]
            merged_cell.value = k[6]
            merged_cell.alignment = alignment
        merged_cell = sheet1["I" +str(j)]
        merged_cell.value = k[7]
        merged_cell.alignment = alignment
        merged_cell = sheet1["J" +str(j)]
        merged_cell.value = k[8]
        merged_cell.alignment = alignment
        merged_cell = sheet1["K" +str(j)]
        merged_cell.value = k[9]
        merged_cell.alignment = alignment
        merged_cell = sheet1["L" +str(j)]
        merged_cell.value = k[10]
        merged_cell.alignment = alignment
        j=j+1
j=int(3)
for i in towns2:#I==key
    merged_cell = sheet2["A" +str(j)]
    merged_cell.value = i
    merged_cell.alignment = alignment
    merged_cell = sheet2["M" +str(j)]
    merged_cell.value = towns2[i][1][0]
    merged_cell.alignment = alignment
    merged_cell = sheet2["N" +str(j)]
    merged_cell.value = towns2[i][1][1]
    merged_cell.alignment = alignment
    merged_cell = sheet2["O" +str(j)]
    merged_cell.value = towns2[i][1][2]
    merged_cell.alignment = alignment
    sheet2.merge_cells("A" + str(j) + ":A" + str(j + int(towns2[i][1][0]-1)))
    sheet2.merge_cells("M" + str(j) + ":M" + str(j + int(towns2[i][1][0]-1)))
    sheet2.merge_cells("N" + str(j) + ":N" + str(j + int(towns2[i][1][0]-1)))
    sheet2.merge_cells("O" + str(j) + ":O" + str(j + int(towns2[i][1][0]-1)))
    for k in towns2[i][0]:#k建案
        merged_cell = sheet2["B" +str(j)]
        merged_cell.value = k[0]
        merged_cell.alignment = alignment
        if "起造人" not in k[1]:
            merged_cell = sheet2["C" +str(j)]
            merged_cell.value = k[1]
            merged_cell.alignment = alignment
        else:
            sheet2.merge_cells("B" + str(j) + ":C" + str(j))
        merged_cell = sheet2["D" +str(j)]
        merged_cell.value = k[2]
        merged_cell.alignment = alignment
        merged_cell = sheet2["E" +str(j)]
        merged_cell.value = k[3]
        merged_cell.alignment = alignment
        if "監造人" in k[4]:
            sheet2.merge_cells("E" + str(j) + ":F" + str(j))
        else:
            merged_cell = sheet2["F" +str(j)]
            merged_cell.value = k[4]
            merged_cell.alignment = alignment
        if "營造廠" in k[5]:
            sheet2.merge_cells("G" + str(j) + ":H" + str(j))
            merged_cell = sheet2["G" +str(j)]
            merged_cell.value = "無"
            merged_cell.alignment = alignment
        else:
            merged_cell = sheet2["G" +str(j)]
            merged_cell.value = k[5]
            merged_cell.alignment = alignment
            merged_cell = sheet2["H" +str(j)]
            merged_cell.value = k[6]
            merged_cell.alignment = alignment
        merged_cell = sheet2["I" +str(j)]
        merged_cell.value = k[7]
        merged_cell.alignment = alignment
        merged_cell = sheet2["J" +str(j)]
        merged_cell.value = k[8]
        merged_cell.alignment = alignment
        merged_cell = sheet2["K" +str(j)]
        merged_cell.value = k[9]
        merged_cell.alignment = alignment
        merged_cell = sheet2["L" +str(j)]
        merged_cell.value = k[10]
        merged_cell.alignment = alignment
        j=j+1
workbook.save("彰化縣~"+str(ee[0])+"年"+str(ee[1])+"月 執照"+".xlsx")